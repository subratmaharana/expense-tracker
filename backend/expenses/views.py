from datetime import datetime, timedelta
from django.utils import timezone
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import inch

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

from django.shortcuts import render , redirect
from django.db.models import Sum ,Count,Q
from django.db.models.functions import TruncMonth
from .forms import IncomeSourceForm , BudgetForm
from django.contrib.auth.decorators import login_required
from .models import Category , Transaction ,IncomeSource

from datetime import datetime
from .models import Budget


def home(request):
    return render(request, "home.html")

@login_required
def dashboard(request):

    # ----------------------------
    # Dashboard Filter
    # ----------------------------

    selected_filter = request.GET.get("filter", "all")

    today = timezone.now().date()

    expense_queryset = Transaction.objects.filter(
        user=request.user,
        transaction_type="Expense"
    )

    income_queryset = IncomeSource.objects.filter(
        user=request.user
    )

    if selected_filter == "today":

        expense_queryset = expense_queryset.filter(
            date=today
        )

        income_queryset = income_queryset.filter(
            date=today
        )
    
    elif selected_filter == "week":

        start_week = today - timedelta(days=today.weekday())

        expense_queryset = expense_queryset.filter(
            date__gte=start_week
        )

        income_queryset = income_queryset.filter(
            date__gte=start_week
        )

    elif selected_filter == "month":

        expense_queryset = expense_queryset.filter(
            date__month=today.month,
            date__year=today.year
        )

        income_queryset = income_queryset.filter(
            date__month=today.month,
            date__year=today.year
        )

    elif selected_filter == "last_month":

        if today.month == 1:
            month = 12
            year = today.year - 1
        else:
            month = today.month - 1
            year = today.year

        expense_queryset = expense_queryset.filter(
            date__month=month,
            date__year=year
        )

        income_queryset = income_queryset.filter(
            date__month=month,
            date__year=year
        )

    elif selected_filter == "year":

        expense_queryset = expense_queryset.filter(
            date__year=today.year
        )

        income_queryset = income_queryset.filter(
            date__year=today.year
        )

    # ----------------------------
    # Summary Cards
    # ----------------------------

    total_income = income_queryset.aggregate(
        total=Sum("amount")
    )["total"] or 0

    total_expense = expense_queryset.aggregate(
        total=Sum("amount")
    )["total"] or 0

    remaining_balance = total_income - total_expense

    total_transactions = expense_queryset.count()

    recent_transactions = expense_queryset.order_by("-date")[:5]

    # ----------------------------
    # Budget
    # ----------------------------

    current_month = datetime.now().month
    current_year = datetime.now().year

    budget = Budget.objects.filter(
        user=request.user,
        month=current_month,
        year=current_year
    ).first()

    budget_amount = budget.amount if budget else 0

    remaining_budget = budget_amount - total_expense

    budget_used = 0

    if budget_amount > 0:

        budget_used = (total_expense / budget_amount) * 100

    if budget_used > 100:

        budget_used = 100

    budget_status = "success"
    budget_message = "You are within your budget."

    if budget_used >= 70 and budget_used < 90:

        budget_status = "warning"

        budget_message = (
            "Warning! You have used over 70% of your budget."
        )

    elif budget_used >= 90:

        budget_status = "danger"

        budget_message = (
            "Alert! Your budget is almost exhausted."
        )

    # ----------------------------
    # Pie Chart
    # ----------------------------

    category_expenses = (
        expense_queryset
        .values("category__name")
        .annotate(total=Sum("amount"))
    )

    chart_labels = []
    chart_data = []

    for item in category_expenses:

        chart_labels.append(item["category__name"])

        chart_data.append(float(item["total"]))

    # ----------------------------
    # Continue in Part 2
    # ----------------------------
    # ----------------------------
    # Monthly Expense Bar Chart
    # ----------------------------

    monthly_expenses = (
        expense_queryset
        .annotate(month=TruncMonth("date"))
        .values("month")
        .annotate(total=Sum("amount"))
        .order_by("month")
    )

    bar_labels = []
    bar_data = []

    for item in monthly_expenses:
        bar_labels.append(item["month"].strftime("%b"))
        bar_data.append(float(item["total"]))

    # ----------------------------
    # Income vs Expense Line Chart
    # ----------------------------

    monthly_income = (
        income_queryset
        .annotate(month=TruncMonth("date"))
        .values("month")
        .annotate(total=Sum("amount"))
        .order_by("month")
    )

    income_dict = {}
    expense_dict = {}

    for item in monthly_income:
        income_dict[item["month"].strftime("%b")] = float(item["total"])

    for item in monthly_expenses:
        expense_dict[item["month"].strftime("%b")] = float(item["total"])

    months = [
        "Jan", "Feb", "Mar", "Apr", "May", "Jun",
        "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
    ]

    line_labels = []
    income_data = []
    expense_data = []

    for month in months:
        if month in income_dict or month in expense_dict:
            line_labels.append(month)
            income_data.append(income_dict.get(month, 0))
            expense_data.append(expense_dict.get(month, 0))

    # ----------------------------
    # Spending Analytics
    # ----------------------------

    top_category = (
        expense_queryset
        .values("category__name")
        .annotate(total=Sum("amount"))
        .order_by("-total")
        .first()
    )

    highest_expense = (
        expense_queryset
        .order_by("-amount")
        .first()
    )

    average_daily_spending = 0

    if total_expense > 0:
        average_daily_spending = round(total_expense / 30, 2)

    monthly_transactions = expense_queryset.filter(
        date__month=datetime.now().month,
        date__year=datetime.now().year
    ).count()

    # ----------------------------
    # Context
    # ----------------------------

    context = {
        "selected_filter": selected_filter,

        "total_income": total_income,
        "total_expense": total_expense,
        "remaining_balance": remaining_balance,
        "total_transactions": total_transactions,
        "recent_transactions": recent_transactions,

        "chart_labels": chart_labels,
        "chart_data": chart_data,

        "bar_labels": bar_labels,
        "bar_data": bar_data,

        "line_labels": line_labels,
        "income_data": income_data,
        "expense_data": expense_data,

        "budget_amount": budget_amount,
        "remaining_budget": remaining_budget,
        "budget_used": budget_used,
        "budget_status": budget_status,
        "budget_message": budget_message,

        "top_category": top_category,
        "highest_expense": highest_expense,
        "average_daily_spending": average_daily_spending,
        "monthly_transactions": monthly_transactions,
    }

    return render(
        request,
        "expenses/dashboard.html",
        context,
    )
@login_required
def reports(request):
    current_month = datetime.now().month
    current_year = datetime.now().year

    total_income = IncomeSource.objects.filter(
        user=request.user,
        date__month=current_month,
        date__year=current_year
    ).aggregate(total=Sum("amount"))["total"] or 0

    total_expense = Transaction.objects.filter(
        user=request.user,
        transaction_type="Expense",
        date__month=current_month,
        date__year=current_year

    ).aggregate(total=Sum("amount"))["total"] or 0

    remaining_balance = total_income - total_expense

    budget = Budget.objects.filter(
        user=request.user,
        month=datetime.now().month,
        year=datetime.now().year
    ).first()

    budget_amount = budget.amount if budget else 0

    budget_used = 0

    if budget_amount > 0:
        budget_used = round((total_expense / budget_amount) * 100, 2)

    top_category = (
        Transaction.objects.filter(
            user=request.user,
            transaction_type="Expense" ,
            date__month=current_month,
            date__year=current_year
        )
        .values("category__name")
        .annotate(total=Sum("amount"))
        .order_by("-total")
        .first()
    )

    recent_transactions = (
        Transaction.objects.filter(
            user=request.user,
            date__month=current_month,
            date__year=current_year
        )
        .order_by("-date")[:10]
    )

    context = {
        "total_income": total_income,
        "total_expense": total_expense,
        "remaining_balance": remaining_balance,
        "budget_amount": budget_amount,
        "budget_used": budget_used,
        "top_category": top_category,
        "recent_transactions": recent_transactions,
    }

    return render(
        request,
        "expenses/reports.html",
        context,
    )


@login_required
def view_category(request):
    categories = Category.objects.all()

    return render(request, "expenses/view_categories.html", {
        "categories": categories
    })


@login_required
def add_expense(request):

    categories = Category.objects.all()
    income_sources = IncomeSource.objects.filter(user=request.user)

    if request.method == "POST":

        Transaction.objects.create(
            user=request.user,
            title=request.POST.get("title"),
            amount=request.POST.get("amount"),
            transaction_type="Expense",
            category=Category.objects.get(id=request.POST.get("category")),
            paid_from=IncomeSource.objects.get(id=request.POST.get("paid_from")),
            date=request.POST.get("date"),
            description=request.POST.get("description"),
        )

        return redirect("expense_history")

    return render(request, "expenses/add_expense.html", {
        "categories": categories,
        "income_sources": income_sources,
    })

from django.db.models import Q

@login_required
def expense_history(request):

    search=request.GET.get("search","")

    expenses=Transaction.objects.filter(
        user=request.user,
        transaction_type="Expense"
    )

    if search:

        expenses=expenses.filter(

            Q(title__icontains=search)|
            Q(category__name__icontains=search)|
            Q(description__icontains=search)

        )

    expenses=expenses.order_by("-date")

    return render(
        request,
        "expenses/expense_history.html",
        {
            "expenses":expenses,
            "search":search
        }
    )

@login_required
def edit_expense(request, id):

    expense = Transaction.objects.get(
        id=id,
        user=request.user
    )

    categories = Category.objects.all()

    if request.method == "POST":

        expense.title = request.POST.get("title")
        expense.amount = request.POST.get("amount")
        expense.transaction_type = request.POST.get("transaction_type")
        expense.category = Category.objects.get(
            id=request.POST.get("category")
        )
        expense.date = request.POST.get("date")
        expense.description = request.POST.get("description")

        expense.save()

        return redirect("expense_history")

    return render(request, "expenses/edit_expense.html", {
        "expense": expense,
        "categories": categories
    })


@login_required
def delete_expense(request, id):

    expense = Transaction.objects.get(
        id=id,
        user=request.user
    )

    expense.delete()

    return redirect("expense_history")


@login_required
def income_list(request):

    incomes = IncomeSource.objects.filter(user=request.user)

    total_income = sum(income.amount for income in incomes)

    return render(request, "expenses/income_list.html", {
        "incomes": incomes,
        "total_income": total_income,
    })

@login_required
def add_income(request):

    if request.method == "POST":

        form = IncomeSourceForm(request.POST)

        if form.is_valid():

            income = form.save(commit=False)
            income.user = request.user
            income.save()

            return redirect("income_list")

    else:

        form = IncomeSourceForm()

    return render(request, "expenses/add_income.html", {
        "form": form
    })

@login_required
def set_budget(request):

    month=datetime.now().month
    year=datetime.now().year
    budget,created=Budget.objects.get_or_create(
        user=request.user,
        month=month,
        year=year,

        defaults={
            "amount":0
        }
    )
    if request.method=="POST":

        form=BudgetForm(request.POST,instance=budget)

        if form.is_valid():

            form.save()

            return redirect("dashboard")

    else:
        form=BudgetForm(instance=budget)
    return render(request,"expenses/set_budget.html",{

        "form":form

    })

@login_required
def export_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"]='attachment; filename="Expense_Report.pdf"'

    doc=SimpleDocTemplate(response,rightMargin=25,leftMargin=25,topMargin=25,bottomMargin=25)
    styles=getSampleStyleSheet()
    elements=[]

    header=Table([["Expense Tracker Pro"]],colWidths=[7*inch])
    header.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#1565C0")),
        ("TEXTCOLOR",(0,0),(-1,-1),colors.white),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),22),
        ("TOPPADDING",(0,0),(-1,-1),14),
        ("BOTTOMPADDING",(0,0),(-1,-1),14),
    ]))
    elements.append(header)
    elements.append(Spacer(1,15))

    title=styles["Heading2"]
    title.alignment=TA_CENTER
    title.textColor=colors.HexColor("#1565C0")
    elements.append(Paragraph("Monthly Expense Report",title))
    elements.append(Spacer(1,15))

    elements.append(Paragraph(f"<b>User:</b> {request.user.username}",styles["Normal"]))
    elements.append(Paragraph(f"<b>Date:</b> {datetime.now().strftime('%d %B %Y')}",styles["Normal"]))
    elements.append(Spacer(1,15))

    total_income=Transaction.objects.filter(user=request.user,transaction_type="Income").aggregate(total=Sum("amount"))["total"] or 0
    total_expense=Transaction.objects.filter(user=request.user,transaction_type="Expense").aggregate(total=Sum("amount"))["total"] or 0
    remaining_balance=total_income-total_expense

    now=datetime.now()
    budget=Budget.objects.filter(user=request.user,month=now.month,year=now.year).first()
    budget_amount=budget.amount if budget else 0

    summary=[
        ["Income","Expense","Balance","Budget"],
        [f"₹{total_income}",f"₹{total_expense}",f"₹{remaining_balance}",f"₹{budget_amount}"]
    ]
    summary_table=Table(summary,colWidths=[1.7*inch]*4)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1565C0")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("BACKGROUND",(0,1),(-1,1),colors.HexColor("#E3F2FD")),
        ("GRID",(0,0),(-1,-1),1,colors.grey),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("TOPPADDING",(0,0),(-1,-1),8),
        ("BOTTOMPADDING",(0,0),(-1,-1),8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1,20))

    data=[["Title","Category","Paid From","Amount","Date"]]
    for expense in Transaction.objects.filter(user=request.user).order_by("-date"):
        data.append([
            expense.title,
            expense.category.name,
            expense.paid_from.source if expense.paid_from else "-",
            f"₹{expense.amount}",
            expense.date.strftime("%d-%m-%Y")
        ])

    table=Table(data,colWidths=[1.8*inch,1.3*inch,1.4*inch,1*inch,1.2*inch])
    style=[
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#263238")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("GRID",(0,0),(-1,-1),0.5,colors.grey),
    ]
    for i in range(1,len(data)):
        style.append(("BACKGROUND",(0,i),(-1,i),colors.whitesmoke if i%2 else colors.beige))
    table.setStyle(TableStyle(style))
    elements.append(table)
    elements.append(Spacer(1,15))

    foot=styles["Italic"]
    foot.alignment=TA_CENTER
    elements.append(Paragraph("Generated by Expense Tracker Pro",foot))

    doc.build(elements)
    return response

@login_required
def export_excel(request):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Expense Report"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Expense_Report.xlsx"'
    )

    # -----------------------------
    # Title
    # -----------------------------

    sheet.merge_cells("A1:E1")

    cell = sheet["A1"]

    cell.value = "Expense Tracker Pro"

    cell.font = Font(size=18, bold=True, color="FFFFFF")

    cell.alignment = Alignment(horizontal="center")

    cell.fill = PatternFill(
        start_color="1565C0",
        end_color="1565C0",
        fill_type="solid",
    )

    # -----------------------------
    # User Details
    # -----------------------------

    sheet["A3"] = "User"
    sheet["B3"] = request.user.username

    sheet["A4"] = "Generated On"
    sheet["B4"] = datetime.now().strftime("%d-%m-%Y %I:%M %p")

    # -----------------------------
    # Summary
    # -----------------------------

    total_income = Transaction.objects.filter(
        user=request.user,
        transaction_type="Income"
    ).aggregate(total=Sum("amount"))["total"] or 0

    total_expense = Transaction.objects.filter(
        user=request.user,
        transaction_type="Expense"
    ).aggregate(total=Sum("amount"))["total"] or 0

    balance = total_income - total_expense

    current_month = datetime.now().month
    current_year = datetime.now().year

    budget = Budget.objects.filter(
        user=request.user,
        month=current_month,
        year=current_year
    ).first()

    budget_amount = budget.amount if budget else 0

    sheet["A6"] = "Total Income"
    sheet["B6"] = total_income

    sheet["A7"] = "Total Expense"
    sheet["B7"] = total_expense

    sheet["A8"] = "Remaining Balance"
    sheet["B8"] = balance

    sheet["A9"] = "Monthly Budget"
    sheet["B9"] = budget_amount

    # -----------------------------
    # Table Heading
    # -----------------------------

    row = 11

    headings = [
        "Title",
        "Category",
        "Paid From",
        "Type",
        "Amount",
        "Date",
    ]

    for col_num, heading in enumerate(headings, 1):

        cell = sheet.cell(row=row, column=col_num)

        cell.value = heading

        cell.font = Font(bold=True, color="FFFFFF")

        cell.fill = PatternFill(
            start_color="263238",
            end_color="263238",
            fill_type="solid",
        )

        cell.alignment = Alignment(horizontal="center")

    # -----------------------------
    # Transactions
    # -----------------------------

    row = 12

    transactions = Transaction.objects.filter(
        user=request.user
    ).order_by("-date")

    for transaction in transactions:

        sheet.cell(row=row, column=1).value = transaction.title

        sheet.cell(row=row, column=2).value = transaction.category.name

        sheet.cell(row=row, column=3).value = (
            transaction.paid_from.source
            if transaction.paid_from
            else "-"
        )

        sheet.cell(row=row, column=4).value = transaction.transaction_type

        sheet.cell(row=row, column=5).value = float(transaction.amount)

        sheet.cell(row=row, column=6).value = (
            transaction.date.strftime("%d-%m-%Y")
        )

        row += 1

    # -----------------------------
    # Column Width
    # -----------------------------

    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 15
    sheet.column_dimensions["E"].width = 15
    sheet.column_dimensions["F"].width = 18

    workbook.save(response)

    return response
